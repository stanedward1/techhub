import json
import os
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.database import get_db
from app.deps import require_teacher, get_current_user
from app.models import (
    Classroom,
    Communication,
    Exam,
    ImportHistory,
    Leave,
    Point,
    Resource,
    Score,
    Seat,
    Student,
    StudentProfileTag,
    User,
    WeeklyReport,
)
from app.audit import (
    attach_student,
    serialize_list_with_students,
    audit,
    student_name,
    active_student_id_query,
    active_classroom_id_query,
)
from app.config import settings
from app.utils import safe_filename, to_dict
from app.permissions import (
    get_teacher_class_ids,
    is_student_in_teacher_classes,
    apply_student_class_filter,
    ensure_student_operable,
    ensure_class_operable,
)
import uuid

router = APIRouter(tags=["教师工作台"])

dep = require_teacher


# ---------------- 成绩 ----------------
@router.get("/api/scores")
def list_scores(
    page: int = 1,
    page_size: int = 20,
    student_id: int | None = None,
    class_id: int | None = None,
    subject: str = "",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(Score)
    # 排除退学学生
    q = q.filter(Score.student_id.in_(active_student_id_query(db)))
    # 教师只能查看自己负责班级的学生成绩
    if user.role != "admin":
        class_ids = get_teacher_class_ids(db, user.id)
        if class_ids:
            student_ids = [s.id for s in db.query(Student).filter(Student.class_id.in_(class_ids)).all()]
            q = q.filter(Score.student_id.in_(student_ids))
        else:
            return {"items": [], "total": 0}
    if class_id:
        # 按班级过滤 + 教师权限校验
        q, denied = apply_student_class_filter(db, user, q, class_id, Score)
        if denied:
            return {"items": [], "total": 0}
    if student_id:
        # 教师只能查看自己班级学生的成绩
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
            return {"items": [], "total": 0}
        q = q.filter(Score.student_id == student_id)
    if subject:
        q = q.filter(Score.subject == subject)
    total = q.count()
    rows = q.order_by(Score.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    items = serialize_list_with_students(db, rows)
    return {"items": items, "total": total}


@router.post("/api/scores")
def create_score(payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not payload.get("student_id") or payload.get("score") is None:
        raise HTTPException(status_code=400, detail="请选择学生并填写成绩")
    # 退学/毕业限制
    ensure_student_operable(db, payload["student_id"])
    # 教师只能为自己班级的学生创建成绩
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, payload["student_id"]):
        raise HTTPException(status_code=403, detail="无权为该学生创建成绩")
    s = Score(
        student_id=payload["student_id"],
        subject=payload.get("subject", "未分类"),
        score=payload["score"],
        exam_name=payload.get("exam_name"),
    )
    db.add(s)
    audit(db, user, "create_score", target=f"新增成绩-{student_name(db, s.student_id)}", student_id=s.student_id, detail=f"科目：{s.subject}；分数：{s.score}；考试：{s.exam_name or '日常测验'}")
    db.commit()
    db.refresh(s)
    return attach_student(db, to_dict(s), s.student_id)


@router.put("/api/scores/{score_id}")
def update_score(score_id: int, payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    s = db.get(Score, score_id)
    if not s:
        raise HTTPException(status_code=404, detail="记录不存在")
    # 退学/毕业限制
    ensure_student_operable(db, s.student_id)
    # 教师只能修改自己班级学生的成绩
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, s.student_id):
        raise HTTPException(status_code=403, detail="无权修改该成绩")
    for f in ("student_id", "subject", "score", "exam_name"):
        if f in payload and payload[f] is not None:
            # 教师不能将成绩转移到其他班级的学生
            if f == "student_id" and user.role != "admin":
                if not is_student_in_teacher_classes(db, user.id, payload["student_id"]):
                    raise HTTPException(status_code=403, detail="无权将成绩转移到该学生")
            if f == "student_id" and payload["student_id"] != s.student_id:
                ensure_student_operable(db, payload["student_id"])
            setattr(s, f, payload[f])
    audit(db, user, "update_score", target=f"成绩#{score_id}-{student_name(db, s.student_id)}", student_id=s.student_id)
    db.commit()
    db.refresh(s)
    return attach_student(db, to_dict(s), s.student_id)


@router.delete("/api/scores/{score_id}")
def delete_score(score_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    s = db.get(Score, score_id)
    if s:
        # 退学/毕业限制
        ensure_student_operable(db, s.student_id)
        # 教师只能删除自己班级学生的成绩
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, s.student_id):
            raise HTTPException(status_code=403, detail="无权删除该成绩")
        db.delete(s)
        audit(db, user, "delete_score", target=f"成绩#{score_id}-{student_name(db, s.student_id)}", student_id=s.student_id)
        db.commit()
    return {"ok": True}


@router.get("/api/scores/export")
def export_scores(student_id: int | None = None, class_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(Score)
    # 排除退学学生
    q = q.filter(Score.student_id.in_(active_student_id_query(db)))
    # 教师只能导出自己负责班级的学生成绩
    if user.role != "admin":
        class_ids = get_teacher_class_ids(db, user.id)
        if class_ids:
            student_ids = [s.id for s in db.query(Student).filter(Student.class_id.in_(class_ids)).all()]
            q = q.filter(Score.student_id.in_(student_ids))
        else:
            student_ids = []
    if class_id:
        q, denied = apply_student_class_filter(db, user, q, class_id, Score)
        if denied:
            student_ids = []
    if student_id:
        # 教师只能导出自己班级学生的成绩
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
            student_ids = []
        q = q.filter(Score.student_id == student_id)
    rows = q.order_by(Score.student_id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩单"
    ws.append(["学号", "姓名", "科目", "成绩", "考试名称"])
    for s in rows:
        stu = db.get(Student, s.student_id)
        ws.append([stu.student_no if stu else "", stu.name if stu else "", s.subject, s.score, s.exam_name or ""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=scores.xlsx"},
    )


# ---------------- 请假/考勤 ----------------
@router.get("/api/leaves")
def list_leaves(
    page: int = 1,
    page_size: int = 20,
    student_id: int | None = None,
    class_id: int | None = None,
    status: str = "",
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(Leave)
    # 排除退学学生
    q = q.filter(Leave.student_id.in_(active_student_id_query(db)))
    # 教师只能查看自己负责班级的学生请假
    if user.role != "admin":
        class_ids = get_teacher_class_ids(db, user.id)
        if class_ids:
            student_ids = [s.id for s in db.query(Student).filter(Student.class_id.in_(class_ids)).all()]
            q = q.filter(Leave.student_id.in_(student_ids))
        else:
            return {"items": [], "total": 0}
    if class_id:
        q, denied = apply_student_class_filter(db, user, q, class_id, Leave)
        if denied:
            return {"items": [], "total": 0}
    if student_id:
        # 教师只能查看自己班级学生的请假
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
            return {"items": [], "total": 0}
        q = q.filter(Leave.student_id == student_id)
    if status:
        q = q.filter(Leave.status == status)
    total = q.count()
    rows = q.order_by(Leave.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": serialize_list_with_students(db, rows), "total": total}


@router.post("/api/leaves")
def create_leave(payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not payload.get("student_id"):
        raise HTTPException(status_code=400, detail="请选择学生")
    # 退学/毕业限制
    ensure_student_operable(db, payload["student_id"])
    # 教师只能为自己班级的学生创建请假
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, payload["student_id"]):
        raise HTTPException(status_code=403, detail="无权为该学生创建请假")
    x = Leave(
        student_id=payload["student_id"],
        reason=payload.get("reason"),
        start_date=payload.get("start_date"),
        end_date=payload.get("end_date"),
        status=payload.get("status", "登记"),
        image=payload.get("image"),
    )
    db.add(x)
    audit(db, user, "create_leave", target=f"新增请假-{student_name(db, x.student_id)}", student_id=x.student_id, detail=f"事由：{x.reason or '未填写'}；时间：{x.start_date or ''} ~ {x.end_date or ''}")
    db.commit()
    db.refresh(x)
    return attach_student(db, to_dict(x), x.student_id)


@router.put("/api/leaves/{leave_id}")
def update_leave(leave_id: int, payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    x = db.get(Leave, leave_id)
    if not x:
        raise HTTPException(status_code=404, detail="记录不存在")
    # 退学/毕业限制
    ensure_student_operable(db, x.student_id)
    # 教师只能修改自己班级学生的请假
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, x.student_id):
        raise HTTPException(status_code=403, detail="无权修改该请假")
    for f in ("reason", "start_date", "end_date", "status", "image"):
        if f in payload and payload[f] is not None:
            setattr(x, f, payload[f])
    audit(db, user, "update_leave", target=f"请假#{leave_id}-{student_name(db, x.student_id)}", student_id=x.student_id, detail=f"事由：{x.reason or '未填写'}；时间：{x.start_date or ''} ~ {x.end_date or ''}")
    db.commit()
    db.refresh(x)
    return attach_student(db, to_dict(x), x.student_id)


@router.delete("/api/leaves/{leave_id}")
def delete_leave(leave_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    x = db.get(Leave, leave_id)
    if x:
        # 退学/毕业限制
        ensure_student_operable(db, x.student_id)
        # 教师只能删除自己班级学生的请假
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, x.student_id):
            raise HTTPException(status_code=403, detail="无权删除该请假")
        db.delete(x)
        audit(db, user, "delete_leave", target=f"请假#{leave_id}-{student_name(db, x.student_id)}", student_id=x.student_id, detail=f"事由：{x.reason or '未填写'}")
        db.commit()
    return {"ok": True}


# ---------------- 积分 ----------------
@router.get("/api/points")
def list_points(page: int = 1, page_size: int = 20, student_id: int | None = None, class_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(Point)
    # 排除退学学生
    q = q.filter(Point.student_id.in_(active_student_id_query(db)))
    # 教师只能查看自己负责班级的学生积分
    if user.role != "admin":
        class_ids = get_teacher_class_ids(db, user.id)
        if class_ids:
            student_ids = [s.id for s in db.query(Student).filter(Student.class_id.in_(class_ids)).all()]
            q = q.filter(Point.student_id.in_(student_ids))
        else:
            return {"items": [], "total": 0}
    if class_id:
        q, denied = apply_student_class_filter(db, user, q, class_id, Point)
        if denied:
            return {"items": [], "total": 0}
    if student_id:
        # 教师只能查看自己班级学生的积分
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
            return {"items": [], "total": 0}
        q = q.filter(Point.student_id == student_id)
    total = q.count()
    rows = q.order_by(Point.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": serialize_list_with_students(db, rows), "total": total}


@router.post("/api/points")
def create_point(payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not payload.get("student_id") or payload.get("points") is None:
        raise HTTPException(status_code=400, detail="请选择学生并填写积分")
    # 退学/毕业限制
    ensure_student_operable(db, payload["student_id"])
    # 教师只能为自己班级的学生创建积分
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, payload["student_id"]):
        raise HTTPException(status_code=403, detail="无权为该学生创建积分")
    x = Point(
        student_id=payload["student_id"],
        points=payload["points"],
        reason=payload.get("reason"),
    )
    db.add(x)
    audit(db, user, "create_point", target=f"新增积分-{student_name(db, x.student_id)}", student_id=x.student_id, detail=f"积分：{x.points}分；原因：{x.reason or ''}")
    db.commit()
    db.refresh(x)
    return attach_student(db, to_dict(x), x.student_id)


@router.delete("/api/points/{point_id}")
def delete_point(point_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    x = db.get(Point, point_id)
    if x:
        # 退学/毕业限制
        ensure_student_operable(db, x.student_id)
        # 教师只能删除自己班级学生的积分
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, x.student_id):
            raise HTTPException(status_code=403, detail="无权删除该积分")
        db.delete(x)
        audit(db, user, "delete_point", target=f"积分#{point_id}-{student_name(db, x.student_id)}", student_id=x.student_id)
        db.commit()
    return {"ok": True}


# ---------------- 家校沟通 ----------------
@router.get("/api/communications")
def list_communications(page: int = 1, page_size: int = 20, student_id: int | None = None, class_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(Communication)
    # 排除退学学生
    q = q.filter(Communication.student_id.in_(active_student_id_query(db)))
    # 教师只能查看自己负责班级的学生沟通
    if user.role != "admin":
        class_ids = get_teacher_class_ids(db, user.id)
        if class_ids:
            student_ids = [s.id for s in db.query(Student).filter(Student.class_id.in_(class_ids)).all()]
            q = q.filter(Communication.student_id.in_(student_ids))
        else:
            return {"items": [], "total": 0}
    if class_id:
        q, denied = apply_student_class_filter(db, user, q, class_id, Communication)
        if denied:
            return {"items": [], "total": 0}
    if student_id:
        # 教师只能查看自己班级学生的沟通
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
            return {"items": [], "total": 0}
        q = q.filter(Communication.student_id == student_id)
    total = q.count()
    rows = q.order_by(Communication.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {"items": serialize_list_with_students(db, rows), "total": total}


@router.post("/api/communications")
def create_communication(payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not payload.get("student_id"):
        raise HTTPException(status_code=400, detail="请选择学生")
    # 退学/毕业限制
    ensure_student_operable(db, payload["student_id"])
    # 教师只能为自己班级的学生创建沟通
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, payload["student_id"]):
        raise HTTPException(status_code=403, detail="无权为该学生创建沟通")
    x = Communication(
        student_id=payload["student_id"],
        method=payload.get("method", "电话"),
        content=payload.get("content"),
        feedback=payload.get("feedback"),
    )
    db.add(x)
    audit(db, user, "create_communication", target=f"新增沟通-{student_name(db, x.student_id)}", student_id=x.student_id, detail=f"方式：{x.method or ''}；内容：{(x.content or '')[:50]}")
    db.commit()
    db.refresh(x)
    return attach_student(db, to_dict(x), x.student_id)


@router.delete("/api/communications/{communication_id}")
def delete_communication(communication_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    x = db.get(Communication, communication_id)
    if x:
        # 退学/毕业限制
        ensure_student_operable(db, x.student_id)
        # 教师只能删除自己班级学生的沟通
        if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, x.student_id):
            raise HTTPException(status_code=403, detail="无权删除该沟通")
        db.delete(x)
        audit(db, user, "delete_communication", target=f"沟通#{communication_id}-{student_name(db, x.student_id)}", student_id=x.student_id)
        db.commit()
    return {"ok": True}


# ---------------- 资源 ----------------
@router.get("/api/resources")
def list_resources(keyword: str = "", _=Depends(dep), db: Session = Depends(get_db)):
    q = db.query(Resource)
    if keyword:
        q = q.filter(Resource.name.contains(keyword))
    rows = q.order_by(Resource.id.desc()).all()
    return {"items": [to_dict(x) for x in rows], "total": len(rows)}


@router.post("/api/resources")
def create_resource(payload: dict, _=Depends(dep), db: Session = Depends(get_db)):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="资源名称不能为空")
    x = Resource(
        name=name,
        category=payload.get("category", "其他"),
        filename=payload.get("filename"),
        filepath=payload.get("filepath"),
    )
    db.add(x)
    audit(db, user, "create_resource", target=f"新增资源")
    db.commit()
    db.refresh(x)
    return to_dict(x)


@router.delete("/api/resources/{resource_id}")
def delete_resource(resource_id: int, _=Depends(dep), db: Session = Depends(get_db)):
    x = db.get(Resource, resource_id)
    if x:
        db.delete(x)
        audit(db, user, "delete_resource", target=f"资源#{resource_id}")
        db.commit()
    return {"ok": True}


# ---------------- 试卷（文件上传管理） ----------------
@router.get("/api/exams")
def list_exams(keyword: str = "", _=Depends(dep), db: Session = Depends(get_db)):
    q = db.query(Exam)
    if keyword:
        q = q.filter(Exam.title.contains(keyword))
    rows = q.order_by(Exam.id.desc()).all()
    return {"items": [to_dict(x) for x in rows], "total": len(rows)}


@router.post("/api/exams/upload")
async def upload_exam(
    title: str = "未命名试卷",
    exam_type: str = "单元测验",
    file: UploadFile = File(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """上传试卷文件：支持 .docx / .pdf / .doc 等格式。"""
    if not file:
        raise HTTPException(status_code=400, detail="请选择试卷文件")

    original = file.filename or "exam"
    ext = os.path.splitext(original)[1].lower()
    allowed = {".pdf", ".doc", ".docx"}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"仅支持 PDF/Word 格式，当前文件类型：{ext}")

    name = safe_filename(os.path.splitext(original)[0]) + "_" + uuid.uuid4().hex[:8] + ext
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    dest = os.path.join(settings.UPLOAD_DIR, name)

    size = 0
    chunk_size = 1024 * 1024
    try:
        with open(dest, "wb") as f:
            while True:
                chunk = await file.read(chunk_size)
                if not chunk:
                    break
                size += len(chunk)
                if size > settings.MAX_UPLOAD_SIZE:
                    f.close()
                    os.remove(dest)
                    raise HTTPException(status_code=413, detail=f"文件过大，最大 {settings.MAX_UPLOAD_SIZE // (1024*1024)}MB")
                f.write(chunk)
    except HTTPException:
        raise
    except Exception:
        if os.path.exists(dest):
            os.remove(dest)
        raise

    x = Exam(
        title=title.strip() or "未命名试卷",
        exam_type=exam_type,
        filename=original,
        filepath=name,
        filesize=size,
        filetype=ext,
    )
    db.add(x)
    audit(db, user, "upload_exam", target=f"试卷上传-{original}")
    db.commit()
    db.refresh(x)
    return to_dict(x)


@router.put("/api/exams/{exam_id}")
def update_exam(exam_id: int, payload: dict, _=Depends(dep), db: Session = Depends(get_db)):
    x = db.get(Exam, exam_id)
    if not x:
        raise HTTPException(status_code=404, detail="试卷不存在")
    for f in ("title", "exam_type"):
        if f in payload and payload[f] is not None:
            setattr(x, f, payload[f])
    audit(db, user, "update_exam", target=f"试卷#{exam_id}")
    db.commit()
    db.refresh(x)
    return to_dict(x)


@router.get("/api/exams/{exam_id}/download")
def download_exam(exam_id: int, _=Depends(dep), db: Session = Depends(get_db)):
    x = db.get(Exam, exam_id)
    if not x or not x.filepath:
        raise HTTPException(status_code=404, detail="文件不存在")
    file_path = os.path.join(settings.UPLOAD_DIR, x.filepath)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件不存在")
    return FileResponse(file_path, filename=x.filename or x.filepath, media_type="application/octet-stream")


@router.delete("/api/exams/{exam_id}")
def delete_exam(exam_id: int, _=Depends(dep), db: Session = Depends(get_db)):
    x = db.get(Exam, exam_id)
    if x:
        # 删除关联文件
        if x.filepath:
            fp = os.path.join(settings.UPLOAD_DIR, x.filepath)
            if os.path.exists(fp):
                os.remove(fp)
        db.delete(x)
        audit(db, user, "delete_exam", target=f"试卷#{exam_id}")
        db.commit()
    return {"ok": True}


# ---------------- 座位表 ----------------
@router.get("/api/seats")
def get_seat(class_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # 教师只能查看自己负责班级的座位
    if user.role != "admin":
        from app.permissions import is_teacher_class_owner
        if not is_teacher_class_owner(db, user.id, class_id):
            raise HTTPException(status_code=403, detail="无权查看该班级座位表")
    s = db.query(Seat).filter(Seat.class_id == class_id).first()
    if not s:
        return {"layout": [], "columns": 6}
    return {"layout": json.loads(s.layout) if s.layout else [], "columns": s.columns}


@router.put("/api/seats")
def save_seat(payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    class_id = payload.get("class_id")
    if not class_id:
        raise HTTPException(status_code=400, detail="缺少班级")
    # 毕业限制：毕业班级不可再修改座位表
    ensure_class_operable(db, class_id)
    # 教师只能保存自己负责班级的座位
    if user.role != "admin":
        from app.permissions import is_teacher_class_owner
        if not is_teacher_class_owner(db, user.id, class_id):
            raise HTTPException(status_code=403, detail="无权修改该班级座位表")
    s = db.query(Seat).filter(Seat.class_id == class_id).first()
    if not s:
        s = Seat(class_id=class_id)
        db.add(s)
    s.layout = json.dumps(payload.get("layout", []), ensure_ascii=False)
    s.columns = payload.get("columns", 6)
    audit(db, user, "save_seat", target=f"保存座位表")
    db.commit()
    return {"ok": True}


# ---------------- 数据导入 ----------------
_EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _validate_student_row(row_data: dict, row_num: int) -> list[str]:
    """校验学生导入行数据，返回错误列表。"""
    errors = []
    if not row_data.get("name", "").strip():
        errors.append(f"第{row_num}行：姓名不能为空")
    if not row_data.get("student_no", "").strip():
        errors.append(f"第{row_num}行：学号不能为空")
    if not row_data.get("class_name", "").strip():
        errors.append(f"第{row_num}行：班级不能为空")
    return errors


def _validate_score_row(row_data: dict, row_num: int) -> list[str]:
    """校验成绩导入行数据，返回错误列表。"""
    errors = []
    if not row_data.get("student_no", "").strip():
        errors.append(f"第{row_num}行：学号不能为空")
    if not row_data.get("subject", "").strip():
        errors.append(f"第{row_num}行：科目不能为空")
    try:
        float(row_data.get("score", 0))
    except (ValueError, TypeError):
        errors.append(f"第{row_num}行：成绩必须是数字")
    return errors


@router.get("/api/students/template")
def download_student_template(_=Depends(dep)):
    """下载学生导入模板。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    ws.append(["学号", "姓名", "性别", "班级", "专业", "出生日期", "家长姓名", "家长电话", "学生类型"])
    ws.append(["2024001", "张三", "男", "2024级1班", "计算机", "2008-05-12", "张父", "13800000000", "通学生"])
    # 设置列宽
    for col, w in enumerate([12, 10, 6, 14, 14, 12, 10, 14, 10], 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type=_EXCEL_MIME, headers={"Content-Disposition": "attachment; filename=student_template.xlsx"})


@router.get("/api/scores/template")
def download_score_template(_=Depends(dep)):
    """下载成绩导入模板。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩导入模板"
    ws.append(["学号", "姓名", "科目", "成绩", "考试名称"])
    ws.append(["2024001", "张三", "语文", "85", "期中考试"])
    for col, w in enumerate([12, 10, 10, 8, 14], 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = w
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type=_EXCEL_MIME, headers={"Content-Disposition": "attachment; filename=score_template.xlsx"})


@router.post("/api/students/import")
async def import_students(
    file: UploadFile = File(...),
    user=Depends(dep),
    db: Session = Depends(get_db),
):
    """批量导入学生数据。"""
    ext = os.path.splitext(file.filename or "" )[1].lower()
    if ext not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式")

    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据行")

    all_errors = []
    success = 0
    total = 0

    # 教师只能导入到自己负责的班级
    if user.role != "admin":
        teacher_class_ids = get_teacher_class_ids(db, user.id)
    else:
        teacher_class_ids = None

    # 预加载班级映射：class_name -> class_id
    classrooms = {c.name: c.id for c in db.query(Classroom).all()}
    graduated_class_ids = {
        c.id for c in db.query(Classroom).filter(Classroom.is_graduated.is_(True)).all()
    }

    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue
        total += 1
        data = {
            "student_no": str(row[0] or "").strip(),
            "name": str(row[1] or "").strip(),
            "gender": str(row[2] or "").strip() or "男",
            "class_name": str(row[3] or "").strip(),
            "major": str(row[4] or "").strip(),
            "birth_date": str(row[5] or "").strip(),
            "parent_name": str(row[6] or "").strip(),
            "parent_phone": str(row[7] or "").strip(),
            "student_type": str(row[8] or "").strip() or "day",
        }
        data["student_type"] = data["student_type"] if data["student_type"] in ("day", "boarding") else "day"

        errors = _validate_student_row(data, row_num)
        if errors:
            all_errors.extend(errors)
            continue

        # 检查学号是否重复
        if db.query(Student).filter(Student.student_no == data["student_no"]).first():
            all_errors.append(f"第{row_num}行：学号 {data['student_no']} 已存在")
            continue

        class_id = classrooms.get(data["class_name"])
        if not class_id:
            all_errors.append(f"第{row_num}行：班级「{data['class_name']}」不存在")
            continue

        # 已毕业班级不可再导入学生
        if class_id in graduated_class_ids:
            all_errors.append(f"第{row_num}行：班级「{data['class_name']}」已毕业，无法导入学生")
            continue

        # 教师只能导入到自己负责的班级
        if teacher_class_ids is not None and class_id not in teacher_class_ids:
            all_errors.append(f"第{row_num}行：教师只能导入到自己负责的班级「{data['class_name']}」")
            continue

        try:
            s = Student(
                student_no=data["student_no"],
                name=data["name"],
                gender=data["gender"],
                class_id=class_id,
                major=data["major"],
                birth_date=data["birth_date"],
                parent_name=data["parent_name"],
                parent_phone=data["parent_phone"],
                student_type=data["student_type"],
            )
            db.add(s)
            db.flush()
            # 自动创建学生账号
            from app.security import hash_password
            exists_user = db.query(User).filter(
                User.role == "student", User.class_id == class_id, User.name == data["name"]
            ).first()
            if not exists_user:
                db.add(User(
                    username=data["name"],
                    password_hash=hash_password("123456"),
                    name=data["name"],
                    role="student",
                    class_id=class_id,
                ))
            success += 1
        except Exception as e:
            all_errors.append(f"第{row_num}行：导入失败 - {str(e)}")

    db.commit()
    audit(db, user, "import_students", target=f"{file.filename or ''} 成功{success}条")

    # 记录导入历史
    db.add(ImportHistory(
        import_type="student",
        filename=file.filename or "",
        total_rows=total,
        success_rows=success,
        error_rows=len(all_errors),
        errors=json.dumps(all_errors[:100], ensure_ascii=False),
        user_id=user.id,
    ))
    db.commit()

    return {"success": success, "total": total, "errors": all_errors[:50]}


@router.post("/api/scores/import")
async def import_scores(
    file: UploadFile = File(...),
    user=Depends(dep),
    db: Session = Depends(get_db),
):
    """批量导入成绩数据。"""
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式")

    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有数据行")

    all_errors = []
    success = 0
    total = 0

    # 教师只能导入自己班级学生的成绩
    if user.role != "admin":
        teacher_class_ids = get_teacher_class_ids(db, user.id)
        teacher_student_nos = set()
        if teacher_class_ids:
            for s in db.query(Student).filter(Student.class_id.in_(teacher_class_ids)).all():
                teacher_student_nos.add(s.student_no)
    else:
        teacher_student_nos = None

    # 预加载学号映射
    student_map = {s.student_no: s for s in db.query(Student).all()}

    for row_num, row in enumerate(rows, start=2):
        if not any(row):
            continue
        total += 1
        data = {
            "student_no": str(row[0] or "").strip(),
            "name": str(row[1] or "").strip(),
            "subject": str(row[2] or "").strip(),
            "score": row[3],
            "exam_name": str(row[4] or "").strip(),
        }

        errors = _validate_score_row(data, row_num)
        if errors:
            all_errors.extend(errors)
            continue

        student = student_map.get(data["student_no"])
        if not student:
            all_errors.append(f"第{row_num}行：学号 {data['student_no']} 不存在")
            continue

        # 退学/毕业限制：退学学生或毕业班级学生不可导入成绩
        if student.is_dropped_out:
            all_errors.append(f"第{row_num}行：学生「{data['name']}」已退学，无法导入成绩")
            continue
        if student.class_id:
            cls = db.get(Classroom, student.class_id)
            if cls and cls.is_graduated:
                all_errors.append(f"第{row_num}行：学生「{data['name']}」所在班级已毕业，无法导入成绩")
                continue

        # 教师只能导入自己班级学生的成绩
        if teacher_student_nos is not None and data["student_no"] not in teacher_student_nos:
            all_errors.append(f"第{row_num}行：教师只能导入自己班级学生「{data['name']}」的成绩")
            continue

        try:
            db.add(Score(
                student_id=student.id,
                subject=data["subject"],
                score=float(data["score"]),
                exam_name=data["exam_name"],
            ))
            success += 1
        except Exception as e:
            all_errors.append(f"第{row_num}行：导入失败 - {str(e)}")

    db.commit()
    audit(db, user, "import_scores", target=f"{file.filename or ''} 成功{success}条")

    db.add(ImportHistory(
        import_type="score",
        filename=file.filename or "",
        total_rows=total,
        success_rows=success,
        error_rows=len(all_errors),
        errors=json.dumps(all_errors[:100], ensure_ascii=False),
        user_id=user.id,
    ))
    db.commit()

    return {"success": success, "total": total, "errors": all_errors[:50]}


@router.get("/api/import-history")
def list_import_history(
    import_type: str = "",
    page: int = 1,
    page_size: int = 20,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """查询导入历史记录。"""
    q = db.query(ImportHistory)
    if import_type:
        q = q.filter(ImportHistory.import_type == import_type)
    total = q.count()
    rows = q.order_by(ImportHistory.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    items = []
    for r in rows:
        d = to_dict(r)
        if r.errors:
            try:
                d["error_list"] = json.loads(r.errors)
            except (json.JSONDecodeError, TypeError):
                d["error_list"] = []
        else:
            d["error_list"] = []
        items.append(d)
    return {"items": items, "total": total}


# ---------------- 学生数字画像 ----------------
@router.get("/api/students/{student_id}/profile")
def get_student_profile(student_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """获取学生综合数字画像数据。"""
    student = db.get(Student, student_id)
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    # 教师只能查看自己班级学生的画像
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
        raise HTTPException(status_code=403, detail="无权查看该学生画像")

    # 成绩统计
    scores = db.query(Score).filter(Score.student_id == student_id).order_by(Score.created_at).all()
    score_summary = {
        "total": len(scores),
        "avg": round(sum(s.score for s in scores) / len(scores), 1) if scores else 0,
        "max": max((s.score for s in scores), default=0),
        "min": min((s.score for s in scores), default=0),
        "by_subject": {},
        "trend": [],
    }
    for s in scores:
        score_summary["by_subject"].setdefault(s.subject, []).append({"score": s.score, "exam": s.exam_name or "", "date": str(s.created_at)[:10]})
        score_summary["trend"].append({"subject": s.subject, "score": s.score, "exam": s.exam_name or "", "date": str(s.created_at)[:10]})

    # 积分统计
    points = db.query(Point).filter(Point.student_id == student_id).all()
    point_summary = {
        "total": sum(p.points for p in points),
        "positive": sum(p.points for p in points if p.points > 0),
        "negative": sum(p.points for p in points if p.points < 0),
        "count": len(points),
        "timeline": [{"points": p.points, "reason": p.reason, "date": str(p.created_at)[:10]} for p in sorted(points, key=lambda x: x.created_at, reverse=True)[:20]],
    }

    # 考勤统计
    leaves = db.query(Leave).filter(Leave.student_id == student_id).all()
    leave_summary = {
        "total": len(leaves),
        "recent": [{"reason": l.reason, "start": l.start_date, "end": l.end_date, "status": l.status} for l in leaves[:10]],
    }

    # 表现统计
    from app.models import Performance
    performances = db.query(Performance).filter(Performance.student_id == student_id).all()
    performance_summary = {
        "positive": sum(1 for p in performances if p.ptype == "积极"),
        "negative": sum(1 for p in performances if p.ptype == "消极"),
        "total": len(performances),
        "recent": [{"ptype": p.ptype, "content": p.content, "date": str(p.created_at)[:10]} for p in sorted(performances, key=lambda x: x.created_at, reverse=True)[:10]],
    }

    # 作业统计
    from app.models import Submission, ExcellentWork
    # Submission.student_id 引用 users.id，需通过学生姓名+班级查找对应 user
    student_user = db.query(User).filter(
        User.role == "student", User.name == student.name, User.class_id == student.class_id
    ).first()
    user_id = student_user.id if student_user else None
    if user_id:
        submissions = db.query(Submission).filter(Submission.student_id == user_id).all()
        excellent_ids = {ew.submission_id for ew in db.query(ExcellentWork.submission_id).filter(
            ExcellentWork.submission_id.in_([s.id for s in submissions])
        ).all()}
        excellent_count = sum(1 for s in submissions if s.id in excellent_ids)
        submission_summary = {
            "total": len(submissions),
            "excellent": excellent_count,
            "rate": round(excellent_count / len(submissions) * 100, 1) if submissions else 0,
        }
    else:
        submission_summary = {"total": 0, "excellent": 0, "rate": 0}

    # 标签
    tags = db.query(StudentProfileTag).filter(StudentProfileTag.student_id == student_id).all()
    tag_list = [{"id": t.id, "tag": t.tag, "category": t.category} for t in tags]

    # 五维雷达得分
    radar = {
        "academic": min(100, round(score_summary["avg"] if scores else 50, 1)),
        "moral": min(100, round(50 + point_summary["total"] * 2, 1)) if points else 50,
        "attendance": min(100, round(100 - leave_summary["total"] * 5, 1)),
        "activity": min(100, round(50 + performance_summary["positive"] * 5, 1)),
        "skill": min(100, round(submission_summary["rate"], 1)),
    }

    # 每个评价维度的评价依据说明：数据来源 / 计算方法 / 相关指标
    radar_basis = [
        {
            "key": "academic",
            "name": "学业",
            "score": radar["academic"],
            "source": "学生成绩记录（成绩管理模块）",
            "method": "按全部考试成绩取平均分，无成绩记录时默认 50 分，满分 100",
            "indicators": [
                f"考试记录 {score_summary['total']} 次",
                f"平均分 {score_summary['avg']} 分",
                f"最高 {score_summary['max']} 分 / 最低 {score_summary['min']} 分",
            ],
        },
        {
            "key": "moral",
            "name": "品德",
            "score": radar["moral"],
            "source": "学生积分记录（积分管理模块，含表现联动加分/扣分）",
            "method": "基准 50 分 + 积分总计 × 2（上限 100），无积分记录时默认 50 分",
            "indicators": [
                f"积分总计 {point_summary['total']} 分",
                f"加分 {point_summary['positive']} 分 / 扣分 {point_summary['negative']} 分",
                f"积分记录 {point_summary['count']} 条",
            ],
        },
        {
            "key": "attendance",
            "name": "出勤",
            "score": radar["attendance"],
            "source": "请假/考勤记录（考勤管理模块）",
            "method": "满分 100 分，每请假 1 次扣 5 分，最低 0 分",
            "indicators": [
                f"请假记录 {leave_summary['total']} 次",
            ],
        },
        {
            "key": "activity",
            "name": "活动",
            "score": radar["activity"],
            "source": "学生表现记录（学生表现模块，积极/消极）",
            "method": "基准 50 分 + 积极表现次数 × 5（上限 100），无记录时默认 50 分",
            "indicators": [
                f"表现记录 {performance_summary['total']} 条",
                f"积极 {performance_summary['positive']} 次 / 消极 {performance_summary['negative']} 次",
            ],
        },
        {
            "key": "skill",
            "name": "技能",
            "score": radar["skill"],
            "source": "作业提交与优秀作品（作业平台）",
            "method": "优秀率 = 优秀作品数 ÷ 提交总数 × 100，满分 100",
            "indicators": [
                f"提交 {submission_summary['total']} 次",
                f"优秀作品 {submission_summary['excellent']} 个",
                f"优秀率 {submission_summary['rate']}%",
            ],
        },
    ]

    return {
        "student": _student_out(db, student),
        "radar": radar,
        "radar_basis": radar_basis,
        "score_summary": score_summary,
        "point_summary": point_summary,
        "leave_summary": leave_summary,
        "performance_summary": performance_summary,
        "submission_summary": submission_summary,
        "tags": tag_list,
    }


def _student_out(db: Session, s: Student) -> dict:
    from app.models import Classroom
    d = to_dict(s)
    cls = db.get(Classroom, s.class_id) if s.class_id else None
    d["class_name"] = cls.name if cls else None
    return d


@router.post("/api/students/{student_id}/tags")
def add_student_tag(student_id: int, payload: dict, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # 退学/毕业限制
    ensure_student_operable(db, student_id)
    # 教师只能给自己班级学生添加标签
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
        raise HTTPException(status_code=403, detail="无权为该学生添加标签")
    tag = (payload.get("tag") or "").strip()
    if not tag:
        raise HTTPException(status_code=400, detail="标签不能为空")
    t = StudentProfileTag(
        student_id=student_id,
        tag=tag,
        category=payload.get("category", "自定义"),
    )
    db.add(t)
    audit(db, user, "add_student_tag", target=f"标签-{student_name(db, student_id)}", student_id=student_id)
    db.commit()
    db.refresh(t)
    return to_dict(t)


@router.delete("/api/students/{student_id}/tags/{tag_id}")
def remove_student_tag(student_id: int, tag_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # 退学/毕业限制
    ensure_student_operable(db, student_id)
    # 教师只能删除自己班级学生的标签
    if user.role != "admin" and not is_student_in_teacher_classes(db, user.id, student_id):
        raise HTTPException(status_code=403, detail="无权删除该学生标签")
    t = db.get(StudentProfileTag, tag_id)
    if t and t.student_id == student_id:
        db.delete(t)
        audit(db, user, "remove_student_tag", target=f"标签#{tag_id}-{student_name(db, student_id)}", student_id=student_id)
        db.commit()
    return {"ok": True}


# ---------------- 班级周报 ----------------
@router.get("/api/reports/weekly-data")
def get_weekly_data(class_id: int, week_start: str = "", week_end: str = "", user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # 教师只能查看自己班级的周报数据
    if user.role != "admin":
        from app.permissions import is_teacher_class_owner
        if not is_teacher_class_owner(db, user.id, class_id):
            raise HTTPException(status_code=403, detail="无权查看该班级周报")
    students = db.query(Student).filter(Student.class_id == class_id, Student.is_dropped_out.is_(False)).all()
    student_ids = [s.id for s in students]
    student_map = {s.id: s.name for s in students}  # id → name 映射，避免重复查询
    total_students = len(students)
    if not student_ids:
        return {"error": "该班级无学生"}

    leave_q = db.query(Leave).filter(Leave.student_id.in_(student_ids))
    if week_start and week_end:
        leave_q = leave_q.filter(Leave.start_date >= week_start, Leave.start_date <= week_end)
    leaves = leave_q.all()

    from app.models import Performance
    perf_q = db.query(Performance).filter(Performance.student_id.in_(student_ids))
    if week_start and week_end:
        perf_q = perf_q.filter(Performance.created_at >= week_start, Performance.created_at <= week_end)
    performances = perf_q.all()
    positive_count = sum(1 for p in performances if p.ptype == "积极")
    negative_count = sum(1 for p in performances if p.ptype == "消极")

    from sqlalchemy import func as sqlfunc
    point_ranking = (
        db.query(Point.student_id, sqlfunc.sum(Point.points).label("total"))
        .filter(Point.student_id.in_(student_ids))
        .group_by(Point.student_id)
        .order_by(sqlfunc.sum(Point.points).desc())
        .all()
    )
    top5 = []
    bottom5 = []
    for pr in point_ranking[:5]:
        name = student_map.get(pr.student_id, "")
        if name:
            top5.append({"name": name, "points": pr.total})
    for pr in point_ranking[-5:]:
        name = student_map.get(pr.student_id, "")
        if name:
            bottom5.append({"name": name, "points": pr.total})

    score_q = db.query(Score).filter(Score.student_id.in_(student_ids))
    if week_start and week_end:
        score_q = score_q.filter(Score.created_at >= week_start, Score.created_at <= week_end)
    recent_scores = score_q.all()
    score_avg = round(sum(s.score for s in recent_scores) / len(recent_scores), 1) if recent_scores else 0

    # 批量查询：一次查询获取所有学生的积分/考勤/表现，避免 N+1
    all_points = db.query(Point).filter(Point.student_id.in_(student_ids)).all()
    all_leaves = db.query(Leave).filter(Leave.student_id.in_(student_ids)).all()
    all_perfs = db.query(Performance).filter(Performance.student_id.in_(student_ids)).all()

    # 按 student_id 分组聚合
    point_map = {}
    for p in all_points:
        point_map.setdefault(p.student_id, 0)
        point_map[p.student_id] += p.points
    leave_map = {}
    for l in all_leaves:
        leave_map[l.student_id] = leave_map.get(l.student_id, 0) + 1
    perf_map = {}
    for p in all_perfs:
        perf_map.setdefault(p.student_id, {"positive": 0, "negative": 0})
        if p.ptype == "积极":
            perf_map[p.student_id]["positive"] += 1
        else:
            perf_map[p.student_id]["negative"] += 1

    profile_summaries = []
    for s in students[:10]:
        profile_summaries.append({
            "name": s.name,
            "student_no": s.student_no,
            "points": point_map.get(s.id, 0),
            "leave_count": leave_map.get(s.id, 0),
            "positive": perf_map.get(s.id, {}).get("positive", 0),
            "negative": perf_map.get(s.id, {}).get("negative", 0),
        })

    return {
        "class_id": class_id,
        "total_students": total_students,
        "leave_count": len(leaves),
        "attendance_rate": round((total_students * 5 - len(leaves)) / (total_students * 5) * 100, 1) if total_students else 0,
        "positive_count": positive_count,
        "negative_count": negative_count,
        "score_avg": score_avg,
        "score_count": len(recent_scores),
        "top5": top5,
        "bottom5": bottom5,
        "profile_summaries": profile_summaries,
        "recent_performances": [
            {"student_name": student_map.get(p.student_id, ""), "ptype": p.ptype, "content": p.content}
            for p in sorted(performances, key=lambda x: x.created_at, reverse=True)[:15]
        ],
    }


@router.get("/api/reports")
def list_reports(class_id: int | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(WeeklyReport)
    # 排除毕业班级的周报
    q = q.filter(WeeklyReport.class_id.in_(active_classroom_id_query(db)))
    # 教师只能查看自己班级的周报
    if user.role != "admin":
        class_ids = get_teacher_class_ids(db, user.id)
        if class_ids:
            q = q.filter(WeeklyReport.class_id.in_(class_ids))
        else:
            return {"items": [], "total": 0}
    if class_id:
        # 教师只能查看自己班级的周报
        if user.role != "admin":
            from app.permissions import is_teacher_class_owner
            if not is_teacher_class_owner(db, user.id, class_id):
                raise HTTPException(status_code=403, detail="无权查看该班级周报")
        q = q.filter(WeeklyReport.class_id == class_id)
    rows = q.order_by(WeeklyReport.id.desc()).all()
    items = []
    for r in rows:
        d = to_dict(r)
        cls = db.get(Classroom, r.class_id)
        d["class_name"] = cls.name if cls else ""
        if r.data_snapshot:
            try:
                d["snapshot"] = json.loads(r.data_snapshot)
            except (json.JSONDecodeError, TypeError):
                d["snapshot"] = {}
        items.append(d)
    return {"items": items, "total": len(rows)}


@router.post("/api/reports")
def save_report(payload: dict, user=Depends(dep), db: Session = Depends(get_db)):
    report_id = payload.get("id")
    title = payload.get("title", "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="报告标题不能为空")
    class_id = payload.get("class_id")
    # 教师只能为自己负责的班级生成周报
    if user.role != "admin" and class_id:
        from app.permissions import is_teacher_class_owner
        if not is_teacher_class_owner(db, user.id, class_id):
            raise HTTPException(status_code=403, detail="无权为该班级生成周报")
    if report_id:
        r = db.get(WeeklyReport, report_id)
        if not r:
            raise HTTPException(status_code=404, detail="报告不存在")
        # 毕业限制：毕业班级不可再修改周报
        ensure_class_operable(db, r.class_id)
        # 教师只能修改自己班级的周报
        if user.role != "admin":
            from app.permissions import is_teacher_class_owner
            if not is_teacher_class_owner(db, user.id, r.class_id):
                raise HTTPException(status_code=403, detail="无权修改该周报")
        r.title = title
        r.content = payload.get("content", "")
        r.data_snapshot = json.dumps(payload.get("data_snapshot", {}), ensure_ascii=False)
    else:
        # 毕业限制：毕业班级不可再生成周报
        if class_id:
            ensure_class_operable(db, class_id)
        r = WeeklyReport(
            class_id=class_id,
            title=title,
            week_start=payload.get("week_start"),
            week_end=payload.get("week_end"),
            content=payload.get("content", ""),
            data_snapshot=json.dumps(payload.get("data_snapshot", {}), ensure_ascii=False),
            created_by=user.id,
        )
        db.add(r)
    audit(db, user, "save_report", target=f"保存周报")
    db.commit()
    db.refresh(r)
    return to_dict(r)


@router.delete("/api/reports/{report_id}")
def delete_report(report_id: int, user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    r = db.get(WeeklyReport, report_id)
    if r:
        # 教师只能删除自己班级的周报
        if user.role != "admin":
            from app.permissions import is_teacher_class_owner
            if not is_teacher_class_owner(db, user.id, r.class_id):
                raise HTTPException(status_code=403, detail="无权删除该周报")
        db.delete(r)
        audit(db, user, "delete_report", target=f"周报#{report_id}")
        db.commit()
    return {"ok": True}
